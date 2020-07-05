import ipaddress
import glob
import re
from openpyxl import Workbook

way = glob.glob("C:\\test\config_files\*.txt")
prob = '____________________________________________'
IPadd = []


for file in way:
    with open(file) as newfile:
        for line in newfile:
            IP = re.match("^ ip address (([0-9]{1,3}\.){3}[0-9]{1,3}) (([0-9]{1,3}\.){3}[0-9]{1,3})",line)

            if IP:
                IPadd.append(ipaddress.IPv4Interface(str(IP.group(1)) + "/" + str(IP.group(3))))
wb = Workbook()
ws = wb.active

ws['A1'] = 'Subnet'
ws['B1'] = 'Mask'
ws['C1'] = 'GW'

coll = 2
p = 0
print("%20s" % "Net", "%20s" % "GW")
while p < len(IPadd)-1:
    print("%20s" % IPadd[p].network.network_address, "%10s" % IPadd[p].network.prefixlen ,"%20s" % IPadd[p].ip)
    p +=1
    n = str(IPadd[p].network.network_address)
    m = str(IPadd[p].network.prefixlen)
    g = str(IPadd[p].ip)
    ws.cell(row=coll, column=1, value=n)
    ws.cell(row=coll, column=2, value=m)
    ws.cell(row=coll, column=3, value=g)
    coll += 1

wb.save('Subnet.xlsx')